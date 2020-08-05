"""
openpyxl_demos.py

（openpyxl一：工作簿、工作表、单元格、样式、页面设置）

深入：
自适应列宽设置。


使用：
官网文档：https://openpyxl.readthedocs.io/en/stable/tutorial.html

创建工作簿Workbook对象及属性方法  ：Workbook() / load_workbook();


创建工作表Worksheet对象及属性方法 ：Workbook.create_sheet() /active, 
# Workbook['title'],   #获取工作表单
  Workbook.copy_worksheet(),

# Worksheet.title / sheet_propertiesp.tabColor,
# Worksheet.row_dimensions[行数].height,
# Worksheet.column_dimensions[列名].width,

# Workbook.sheetnames, for sheet in Workbook,  #获取多个工作表单

# Workbook.remove(Worksheet);


单元格cell的相关操作 ：
# Worksheet.cell(), Wroksheet['Cell'],
# Worksheet.cell(row,column,value), Cell.value, Worksheet.append([]),  #写入内容的方式：列号、值、列表。

# Cell.number_format,
# openpyxl.comments.Comment(), Comment.height/width, Cell.comment,

# Cell.iter_rows/iter_cols,
# Worksheet[:], Worksheet.rows / columns / values,
# Worksheet.max_column / max_row,

# Worksheet.merge_cells() / unmerge_cell(),
# Worksheet.column_dimensions.group(),
# Worksheet.row_dimensions.group(),

# Worksheet.insert_cols(),
# Worksheet.insert_rows(),
# Worksheet.delete_cols(),
# Worksheet.delete_rows(),
# Worksheet.move_range(),

# openpyxl.drawing.image.Image(), Worksheet.add_image(),

# openpyxl.worksheet.table.Table()/TableStyleInfo(), 
# Table.tableStyleInfo, 
# Worksheet.add_table(), 

# Worksheet.auto_filter.ref()/add_filter_column()/add_sort_condition();

使用样式styles（单元格样式 和 命名样式NameeStyle/样式模板）: 

# 单元格样式：默认格式设置 以及 单元格样式应用:
# openpyxl.style.Font(),         Cell.font,
# openpyxl.style.PatternFill(),  Cell.fill,
# openpyxl.style.Border(),       Cell.border,
# openpyxl.style.Side(),
# openpyxl.style.Alignment(),    Cell.alignment,
# openpyxl.style.Protection();

    
# 命名样式NameeStyle/样式模板: 
# openpyxl.styles.NamedStyle(), Workbook.add_named_style(), Cell.style;
# 注意：
#    命名样式在首次分配给单元时也将自动注册 如ws['A1'].style = NamedStyle(name='namedstyle')
#    注册后，仅使用名称分配样式 如ws['A2'].style=namedstyle

    

工作表单的页面设置：
# Worksheet.page_setup.orientation / paperSize / fitToHeight / fitToWidth

# Worksheet.print_options.horizontalCentered / verticalCentered
# Worksheet.oddHeader.left.text / size / font / color
# Worksheet.print_title_cols / rows
# Worksheet.print_area


工作簿数据存储 ：Workbook.save()。


"""

# =============================================================================
# #openpyxl
# #openpyxl用于读取/写入Excel文件（ Excel2010 xlsx/xlsm/xltx/xltm文件）
# #
# #
# =============================================================================

########## 探索模块 常用类、方法、属性
import openpyxl
from openpyxl import Workbook                          #新建工作簿类使用
from openpyxl import load_workbook                     #加载工作簿使用
from openpyxl.worksheet.worksheet import Worksheet     #工作表单类

from openpyxl.comments import Comment                  #添加注释内容使用
from openpyxl.drawing.image import Image               #插入图片使用
from openpyxl.worksheet.table import (Table,           #插入表单使用
                                      TableStyleInfo   #表单样式信息类
                                      )

from openpyxl.styles import (NamedStyle,    #使用样式：创建命名样式
                             Font,          #使用样式：字体，用于设置字体大小、颜色、下划线等
                             PatternFill,   #使用样式：图样填充
                             Border,        #使用样式：边框设置
                             Side,          #使用样式：边框类型设置border_style
                             Alignment,     #使用样式：对齐方式
                             Protection,    #使用样式：保护选项
                             colors         #颜色选项
                             )



#################### 创建工作簿Workbook对象及属性方法  ：Workbook() / load_workbook();


#新建工作簿对象
#####
#使用：openpyxl.Workbook((self, write_only=False, iso_dates=False))
#来实例化创建一个工作簿对象，工作簿是文档所有其他部分的容器。
help(Workbook)                   #通过__init__文件定义: from .workbook import Workbook
dir(Workbook)

wb=Workbook()
wb_filename=r'openpyxl_cases\openpyxl_demos_case1.xlsx'

#加载已有工作簿
#####
#使用：openpyxl.load_workbook(filename, read_only=False, keep_vba=False, data_only=False, keep_links=True)
#来加载工作簿。打开给定的文件名并返回工作簿。
#参数filename         : 为打开或类似文件的对象的路径
#参数read_only=False  : 优化阅读，内容不能编辑
#参数keep_vba=False   : 控制是否保留任何Visual Basic元素。如果保留它们，则仍不可编辑。
#参数data_only=False  : 控制带有 公式 的单元格是否具有公式(默认值) 或 上次Excel读取工作表时存储的值
#参数keep_links=True  : 指向外部工作簿的 链接 是否应该保留。默认值为True.
help(openpyxl.load_workbook)    #通过openpyxl/__init__.py文件定义：from openpyxl.reader.excel import load_workbook

#wb=load_workbook(wb_filename)



#################### 创建工作表Worksheet对象及属性方法 ：Workbook.create_sheet() /active, 
#Workbook['title'], Workbook.copy_worksheet(),

#Worksheet.title / sheet_propertiesp.tabColor,
#Worksheet.row_dimensions[行数].height,
#Worksheet.column_dimensions[列名].width,

#自动设置列宽
#####
#使用：for循环遍历得出每列长度后形成字典数据来自动设置每列列宽。

#Workbook.sheetnames, for sheet in Workbook, 

#Workbook.remove(Worksheet);


#创建工作表单、获取活跃工作表单
#####
#使用：Workbook.create_sheet(self, title=None, index=None) 
#来创建工作表，返回Worksheet类对象
#参数title=None       : 指定工作表的标题名称
#参数index=None       : 指定工作表的索引顺序

#####
#使用：Workbook.active 属性 来获取当前活跃的工作表，返回Worksheet类对象
help(Workbook.active)
help(Workbook.create_sheet)

ws=wb.active
ws_jibenqingkuang = wb.create_sheet(title='基本情况',index=1)
ws_dazongjiaoyi = wb.create_sheet(title='大宗交易',index=2)
ws_longhubang = wb.create_sheet(title='龙虎榜',index=3)

#####
#注意：工作表单命名后就可以作为 工作簿的键名来操作，如 ws2=wb['大宗交易']
ws2=wb['大宗交易']

#拷贝工作表单副本
#####
#使用：Workbook.copy_worksheet(self, from_worksheet) 
#来拷贝创建工作表副本
#参数from_worksheet: 指定要复制的工作表单
#注意：如果wb工作簿以只读或者仅写模式打开，则不能赋值ws工作表。
help(Workbook.copy_worksheet)

ws_fuben=wb.copy_worksheet(ws_longhubang)
ws_fuben.title='副本-龙虎榜'

#获取工作表单的title属性
#####
#使用：Worksheet.title 属性 设置title属性来更改名称，如：ws.title='活跃表单'
help(openpyxl)
ws.title='活跃表单'
ws_jibenqingkuang.title='基本资料'

#设置工作表单title背景色
#####
#使用：Worksheet.sheet_properties.tabColor 属性 提供工作表单的颜色代码。默认为白色。
ws.sheet_properties.tabColor='1072BA'

#设置工作表单的行高 和 列宽
#####
#使用：Worksheet.row_dimensions[行数].height 属性 来设置指定行的行高
#使用：Worksheet.column_dimensions[列名].width 属性 来设置指定列的列宽。
ws.row_dimensions[1].height = 20
ws.column_dimensions['B'].width = 20


#自动设置列宽
#####
#使用：for循环遍历得出每列长度后形成字典数据来自动设置每列列宽。
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            """
            首先获取每个单元格中的长度；如果有换行则按单行的长度计算，先分割再计算；
            长度计算中：len('中文')>>>2, len('中文'.encode('utf-8'))>>>6，通过运算，将中文的字节数定义为2；
            字典存储每列的宽度：将cell每列中 列名最为键名，cell长度计算的最大长度作为键值。
            """
            len_cell = max([(len(line.encode('utf-8'))-len(line))/2+len(line) for line in str(cell.value).split('\n')])
            #dims[chr(64+cell.column)] = max((dims.get(chr(64+cell.column), 0), len(str(cell.value))))
            dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len_cell)
for col, value in dims.items():
    """最后通过遍历存储每列的宽度的字典，来设置相关列的宽度"""
    ws.column_dimensions[col].width = value+2 if value+2<=50 else 50



#获取工作表单所有名称属性
#####
#使用：Workbook.sheetnames 属性 来查看工作簿中的所有工作表的名称，返回列表形式。
print(wb.sheetnames)

#####
#使用：for循环直接遍历工作簿wb 来获取每个工作表单ws
for sheet in wb:
    print(sheet.title)

#####
#使用：Workbook.get_sheet_by_name('XXX')  #来获取指定名称的工作表单
ws1 = wb.get_sheet_by_name('活跃表单')
print(ws1.title)

#####
#使用：Workbook.remove(Worksheet) 来删除工作表
#注意：这里的删除工作表参数只能是对象Worksheet，不能写sheetname
wb.remove(ws_fuben)




#################### 单元格cell的相关操作 ：
#Worksheet.cell(), Wroksheet['Cell'],
#Cell.value/ number_format,
#openpyxl.comments.Comment(), Comment.height/width, Cell.comment,

#Cell.iter_rows/iter_cols,
#Worksheet[:], Worksheet.rows/columns/values,

#Worksheet.merge_cells()/unmerge_cell(),
#Worksheet.column_dimensions.group(),
#Worksheet.row_dimensions.group(),

#Worksheet.insert_cols(),
#Worksheet.insert_rows(),
#Worksheet.delete_cols(),
#Worksheet.delete_rows(),
#Worksheet.move_range(),

#openpyxl.drawing.image.Image(), Worksheet.add_image(),

#openpyxl.worksheet.table.Table()/TableStyleInfo(), 
#Table.tableStyleInfo, 
#Worksheet.add_table(), 

#Worksheet.auto_filter.ref()/add_filter_column()/add_sort_condition();



#访问/赋值单个单元格
#####
#使用：Worksheet.cell(self, row, column, value=None)
#来根据给定的坐标返回一个单元格对象,可直接进行赋值操作。
#参数row         : 表示单元格的行索引
#参数column      : 表示单元格的列索引
#参数value=None  : 可以直接给指定单元格赋值操作
#注意：单元格cell的行列索引都是从1开始，而不是python的从0开始。
help(Worksheet.cell)

ws.cell(row=1, column=1,value='第一次添加单元格cell数据')
ws['A2']='A2'

for x in range(3,13):
    for y in range(3,13):
        ws.cell(x,y,1)

#获取单元格的值
#####
#使用：Cell.value 属性 来进行单元格的取值和赋值操作
A2=ws['A2']
A2.value='a2'

#####
#使用：Worksheet.append([]) 来在工作表单底部追加写入列表。
rows = [
        ['Aliens', 2, 3, 4, 5, 6, 7],
        ['Humans', 10, 40, 50, 20, 10, 50]
       ]
for row in rows:
    ws_tjzx.append(row)


#输出数字格式
#####
#使用：Cell.number_format 来输出单元格的数字格式
import datetime
#from openpyxl import Workbook
#wb=Workbook()
#ws=wb.active
ws['B1']=datetime.datetime(2019,11,11)
ws['B1'].number_format
B1=ws['B1']
print(B1.value)

#注释的添加/保存/设置
#####
#使用：openpyxl.comments.Comment(text, author, height=79, width=144)
#来创建一个注释对象。以便接下来通过赋值添加给单元格。
#参数text         : 添加注释内容
#参数author       : 添加作者信息
#参数height=79    : 注释框高度，也可以后期通过Comment.height 属性 来赋值设置高度
#参数width=144    : 注释框宽度，也可以后期通过Comment.hight 属性 来赋值设置宽度

#使用：Cell.comments 属性 来赋值添加该单元格cell的注释
help(openpyxl.comments.Comment)

comment=Comment('这里是添加注释内容...','血皇敖天',height=79,width=144)
comment.height=79
comment.width=300
ws['C22'].comment=comment    #对单元格的comment属性进行赋值操作添加注释对象
ws['I3'].comment=comment     #同一个注释对象 可以 赋值给 多个单元格


#使用公式
#####
#使用：公式使用字符串形式代表，格式如：ws['E1']='=SUM(1,1)'或 ws['E1']='=C1+D1'
ws['C1']=3
ws['D1']=4
ws['E1']='=C1+D1'


#单个及多个单元格
#####
#使用：Worksheet['A2']    键名格式 来表示/获取/赋值cell单元格的行/列/单个单元格
#使用：Worksheet['A15':'D25'] 切片格式 来表示/获取/赋值工作表单中 cell单元格 的范围。如：cell_range = ws['A15':'D25']
#使用：Worksheet['A:D']  切片格式 来表示/获取/赋值工作表单中 列 的范围。如：col_range = ws['A:D']
#使用：Worksheet[5:10]   切片格式 来表示/获取/赋值工作表单中 行 的范围。如：row_range = ws[5:10]
ws['A3']=22
cell_range = ws['A15':'D25']
col_range = ws['A:D']
row_range = ws[5:10]

for row in ws['A15:D25']:  #先遍历出每行
    for cell in row:       #再遍历每行的Cell
        print(cell)


#获取多个单元格数据
#####
#使用：Worksheet.iter_rows(self, min_row=None, max_row=None, min_col=None, max_col=None, values_only=False)
#来 按行从工作表生成单元格。指定迭代范围使用行和列的索引。
#参数min_row=None: 指定单元格的开始行索引
#参数max_row=None: 指定单元格的结束行索引
#参数min_col=None: 指定单元格的开始列索引
#参数max_col=None: 指定单元格的结束列索引
#参数values_only=False： 指定是否只显示单元格的值

#####
#使用：Worksheet.iter_rows() 等同于 Worksheet.iter_cols()
help(Worksheet.iter_rows)
for row in ws.iter_rows(min_row=1,max_row=2,min_col=1,max_col=5):  #先遍历出每行
    for cell in row:                                               #再遍历每一行
        print(cell,cell.value)
#####
#使用：Worksheet.iter_cols() 等同于 Worksheet.iter_rows()
help(Worksheet.iter_cols)
for col in ws.iter_cols(min_row=1,max_row=2,min_col=1,max_col=5, values_only=True):
    for cell in col:
        print(cell)

#访问工作表单所有行列
#####
#使用：Worksheet.rows 属性 来访问工作表单中的所有 行。
#使用：Worksheet.columns 属性 来访问工作表单中的所有 列。
#注意：出于性能原因，Worksheet.columns 属性 在只读模式下不可用。
type(ws.rows)
list(ws.rows)
list(ws.columns)
for row in ws.rows:
    for cell in row:
        print(cell)

for column in ws.columns:
    for cell in column:
        print(cell)

#访问工作表单中所有单元格的值
#####
#使用：Worksheet.values 属性 来获取工作表单中的所有值，按行遍历显示，每行是一个元组tuple。
help(Worksheet.values)
for row in ws.values:
    print(row)

#####
#使用：Worksheet.max_column 属性 来获取工作表最大列数
#使用：Worksheet.max_row 属性 来获取工作表最大行数
ws.max_column
ws.max_row


#合并单元格
#####
#使用：Worksheet.merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None)
#来合并单元格。
#合并单元格时，除左上角以外的所有单元格都将从工作表中删除。
#参数range_string=None : 通过字符串形式指定合并范围，如：ws.merge_cells('A8:D10')
#参数star_row=None     : 指定合并开始行的索引
#参数end_row=None      : 指定合并结束行的索引
#参数start_column=None : 指定合并开始列的索引
#参数end_column=None   : 指定合并结束列的索引

#使用：Worksheet.unmerge_cells()取消合并单元格，参数用法等同于 合并单元格
help(Worksheet.merge_cells)

ws.merge_cells('A8:D10')
#或
ws.merge_cells(start_row=13 ,end_row=15,start_column=1 ,end_column=5)


#折叠行/列
#####
#使用：Worksheet.column_dimensions.group(self,start,end=None,outline_level=1,hidden=False)
#来将工作表单的 列折叠
#参数start           : 指定要分组的第一行或第一列(强制要求)
#参数end=None        : 指定分组的最后一行或最后一列(可选，默认开始)
#参数outline_level=1 : 指定大纲级别
#参数hidden=False    : 表示 组是否应该隐藏在打开的工作簿上

#使用：Worksheet.row_dimensions.group(self,start,end=None,outline_level=1,hidden=False)
#来将工作表单的 行折叠，参数等同 列折叠。
ws.column_dimensions.group('J',end='Q',outline_level=1,hidden=True)
ws.row_dimensions.group(15,25,hidden=True)


#插入行/列
#####
#使用：Worksheet.insert_cols(self, idx, amount=1) 来插入1行，注意参数idx为数字
#使用：Worksheet.insert_rows(self, idx, amount=1) 来插入1列，注意参数idx为数字
help(Worksheet.insert_cols)
ws.insert_rows(6)
ws.insert_cols(4)


#删除行/列
#####
#使用：Worksheet.delete_cols(self, idx, amount=1) 来删除指定行
#使用：Worksheet.delete_rows(self, idx, amount=1) 来删除指定列
help(Worksheet.delete_cols)
help(Worksheet.delete_rows)
#ws.delete_rows(3)
#ws.delete_cols(5,2)


#移动单元格的范围
#####
#使用：Worksheet.move_range(self, cell_range, rows=0, cols=0, translate=False)
#来移动单元格的范围
#参数cell_range   : 指定要移动的单元格的范围。如："D3:F10"
#参数rows         : 指定移动的行数，如：+1向下移动1行，-1向上移动1行
#参数cols         : 指定移动的列数，如：+1向右移动1列，-1向左移动1列
#参数translate=False: 表示现有单元将被覆盖。公式和参考资料将不会更新
help(Worksheet.move_range)
ws.move_range('G10:I13',rows=2,cols=2,translate=True)



#插入图片
#####
#使用：openpyxl.drawing.image.Image(self,img) 来初始化图片
#使用：Worksheet.add_image(self, img, anchor=None) 来插入图片
#参数img: 为经过Image类初始化后的图片对象
#参数anchar: 为cell单元格定位
from openpyxl.drawing.image import Image
help(Image)
help(Worksheet.add_image)

img=Image(r'openpyxl_cases\002592.png')
ws.add_image(img,'A11')



#插入表单Table
#####
from openpyxl import Workbook
from openpyxl.worksheet.table import (Table,           #插入表单使用
                                      TableStyleInfo   #表单样式信息类
                                      )
ws_table=wb.create_sheet(title='Table')                #创建一个工作表

data=[
      ['Apples',10000, 5000,8000, 6000],
      ['Pears',2000,3000,4000,5000],
      ['Bananas',6000,6000,6500,6000],
      ['Oranges',500,300,200,700]
     ]

#####
#使用：Worksheet.append(iterable) 来在当前工作表的底部附加一组值。
#参数iterable: 列表、范围或生成器，或包含要追加的值的dict
help(Worksheet.append)
ws_table.append(['Fruit','2011','2012','2013','2014'])
for row in data:
    ws_table.append(row)

#####
#使用：openpyxl.worksheet.table.Table() 来创建表单对象
help(Table)
tab=Table(displayName='Table1', ref='A1:E5')

#####
#使用：openpyxl.worksheet.table.TableStyleInfo() 来创建表单的样式信息对象。
help(openpyxl.worksheet.table.TableStyleInfo)

style=TableStyleInfo(name='TableStyleMedium9',
                     showFirstColumn=False,
                     showLastColumn=False,
                     showRowStripes=True,
                     showColumnStripes=True)
#####
#使用：Table.tableStyleInfo 属性 来赋值表单的样式信息
tab.tableStyleInfo = style

#####
#使用：Worksheet.add_table(table) 来添加表单。
ws_table.add_table(tab)



#插入过滤器和排序
#####
help(openpyxl.worksheet.filters.AutoFilter)
help(openpyxl.worksheet.filters.AutoFilter.ref)
help(openpyxl.worksheet.filters.AutoFilter.add_filter_column)   #添加指定列的行筛选器。
help(openpyxl.worksheet.filters.AutoFilter.add_sort_condition)  #为指定范围的单元格添加排序条件

ws_guolvqi=wb.create_sheet(title='guolvqi')
data_guolvqi=[
        ['Fruit','Quantity'],
        ['Kiwi',1],
        ['Grape',2],
        ['Apple',3],
        ['Peach',4],
        ['Pomegranate',5],
        ['Pear',6],
        ['Tangerine',7],
        ['Blueberry',8],
        ['Mango',9],
        ['Watermelon',10],
        ['Blackberry',11],
        ['orange',12],
        ['Raspberry',13],
        ['Banana',14]]
for r in data_guolvqi:
    ws_guolvqi.append(r)
    
#####
#使用：Worksheet.auto_filter.ref 属性 赋值 单元格/单元格范围 来 添加过滤器和排序。
#如下：表示 添加过滤器和排序 作用于 单元格范围'A1:B15'
ws_guolvqi.auto_filter.ref='A1:B15'

#####
#使用：Worksheet.auoto_filter.add_filter_column(self, col_id, vals, blank=False)
#来 添加指定列的行筛选器
ws_guolvqi.auto_filter.add_filter_column(0,['Kiwi','Apple','Mango','Pear'])  #添加指定列的行筛选器。

#####
#使用：Worksheet.auto_filter.add_sort_condition(self, ref, descending=False)
#来 为指定范围的单元格添加排序条件。
ws_guolvqi.auto_filter.add_sort_condition('B2:B11')  #为指定范围的单元格添加排序条件






#################### 使用样式styles（单元格样式 和 命名样式NameeStyle/样式模板）: 

#单元格样式：默认格式设置 以及 单元格样式应用:
#openpyxl.style.Font(),         Cell.font,
#openpyxl.style.PatternFill(),  Cell.fill,
#openpyxl.style.Border(),       Cell.border,
#openpyxl.style.Side(),
#openpyxl.style.Alignment(),    Cell.alignment,
#openpyxl.style.Protection();

    
#命名样式NameeStyle/样式模板: 
#openpyxl.styles.NamedStyle(), Workbook.add_named_style(), Cell.style;
#注意：
#    命名样式在首次分配给单元时也将自动注册 如ws['A1'].style = NamedStyle(name='namedstyle')
#    注册后，仅使用名称分配样式 如ws['A2'].style=namedstyle


import openpyxl
from openpyxl import Workbook                          #新建工作簿类使用
from openpyxl import load_workbook                     #加载工作簿使用
from openpyxl.worksheet.worksheet import Worksheet     #工作表单类

from openpyxl.comments import Comment                  #添加注释内容使用
from openpyxl.drawing.image import Image               #插入图片使用
from openpyxl.worksheet.table import (Table,           #插入表单使用
                                      TableStyleInfo   #表单样式信息类
                                      )

from openpyxl.styles import (NamedStyle,    #使用样式：创建命名样式
                             Font,          #使用样式：字体，用于设置字体大小、颜色、下划线等
                             PatternFill,   #使用样式：图样填充
                             Border,        #使用样式：边框设置
                             Side,          #使用样式：边框类型设置border_style
                             Alignment,     #使用样式：对齐方式
                             Protection,    #使用样式：保护选项
                             colors         #颜色选项
                             )

#样式设置的默认值：
font = Font(name='Calibri',     #字体名称 如：微软雅黑、宋体等
            size=11,            #字号
            bold=False,         #粗体
            italic=False,       #斜体
            vertAlign=None,     #纵向对齐
            underline='none',   #下划线（‘doubleAccounting’, ‘single’, ‘double’, ‘singleAccounting’）
            strike=False,       #删除线
            color='FF000000'    #字体颜色
            )

fill = PatternFill(fill_type=None,            #指定填充的类型，支持的有：'solid'等。
                   start_color='ffffffff',    #指定天聪的开始颜色
                   end_color='ff000000'       #指定填充的结束颜色
                   )

border = Border(left=Side(border_style=None, color='FF000000'),        #左边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                right=Side(border_style=None, color='FF000000'),       #右边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                top=Side(border_style=None, color='FF000000'),         #上边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                bottom=Side(border_style=None, color='FF000000'),      #下边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                diagonal=Side(border_style=None, color='FF000000'),    #对角线边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                diagonal_direction=0,                                  #对角线方向
                outline=Side(border_style=None, color='FF000000'),     #外边框线设置，Side类定义 变类型/颜色。'thin'/'thick'
                vertical=Side(border_style=None, color='FF000000'),    #垂直线设置，Side类定义 变类型/颜色。'thin'/'thick'
                horizontal=Side(border_style=None, color='FF000000'),  #水平线设置，Side类定义 变类型/颜色。'thin'/'thick'
                diagonalDown=False,
                start=None,
                end=None
                )

alignment = Alignment(horizontal='general',  #水平对齐('centerContinuous', 'general', 'distributed','left', 'fill', 'center', 'justify', 'right')
                      vertical='bottom',     #垂直对齐（'distributed', 'top', 'center', 'justify', 'bottom'）
                      text_rotation=0,       #文字旋转
                      wrap_text=False,       #自动换行
                      shrink_to_fit=False,   #缩小字体填充
                      mergeCell=None,        #合并单元格
                      indent=0               #缩进
                      )

number_format = 'General'

protection = Protection(locked=True,
                        hidden=False
                        )


#探索模块
help(openpyxl.styles)
dir(openpyxl.styles)

help(openpyxl.styles.Font)
help(openpyxl.styles.PatternFill)
help(openpyxl.styles.Border)
help(openpyxl.styles.Side)
help(openpyxl.styles.Alignment)
help(openpyxl.styles.Protection)



########## #单元格样式：默认格式设置 以及 单元格样式应用:
#openpyxl.style.Font(),         Cell.font,
#openpyxl.style.PatternFill(),  Cell.fill,
#openpyxl.style.Border(),       Cell.border,
#openpyxl.style.Side(),
#openpyxl.style.Alignment(),    Cell.alignment,
#openpyxl.style.Protection();


#字体设置 样式对象
#####
#使用：openpyxl.style.Font(name=None, sz=None, b=None, i=None, charset=None, u=None, 
#strike=None, color=None, scheme=None, family=None, size=None, bold=None, italic=None, 
#strikethrough=None, underline=None, vertAlign=None, outline=None, shadow=None, 
#condense=None, extend=None) 来创建样式对象用于 设置字体大小、颜色、下划线等字体

help(openpyxl.styles.Font)

font = Font(name='Calibri',     #字体名称 如：微软雅黑、宋体等
            size=11,            #字号
            bold=False,         #粗体
            italic=False,       #斜体
            vertAlign=None,     #纵向对齐
            underline='none',   #下划线
            strike=False,       #删除线
            color='FF000000'    #字体颜色
            )

#使用：Cell.font 属性 赋值 openpyxl.sytles.Font样式对象 来添加cell单元格的 字体样式
#注意：字体的颜色参数color 通常是RGB或ARGB十六机制，也可以是openpyxl.styles.colors模块的颜色常量属性。
E12=ws['E12']
E12.font=Font(color=colors.RED)      #设置字体颜色为红色
ft1=Font(color=colors.GREEN)
ws['A1'].font=ft1
ws['A1'].font=Font(bold=True)        #设置单元格字体为粗体

#使用：copy(样式对象) 来复制一个样式对象。注意：需要from copy import copy
from copy import copy
ft2=Font(name='样式2',size=14)
ft3=copy(ft2)
ft3.name

#注意：样式对象 也可以 直接赋值给行/列，来设置整行/整列的样式。
#但是官网提示注意，这仅适用于关闭文件后在Excel中创建的单元格。/已经存在的单元格。（不适用）
#注意：如果要将样式应用于整个行和列，则必须自己将样式应用于每个单元格。
#注意：
#col_F=ws.column_dimensions['I']
#col_F.font=Font(color=colors.YELLOW)
#row_3=ws.row_dimensions[3]
#row_3.font=Font(color=colors.GREEN)



#图样填充 样式对象
#####
#使用：openpyxl.style.PatternFill(fill_type=None, start_color=None, end_color=None) 
#来创建样式对象用于 进行图样填充、渐变色等 
#参数fill_type：如果不指定则后面设置不起作用。支持的值如下：
#{'darkHorizontal', 'gray0625', 'darkGrid', 'darkGray', 'darkUp', 'solid', 'darkVertical', 
#'gray125', 'darkDown', 'lightDown', 'mediumGray', 'darkTrellis', 'lightHorizontal',
# 'lightTrellis', 'lightUp', 'lightGrid', 'lightGray', 'lightVertical'}
help(openpyxl.styles.PatternFill)

fill = PatternFill(fill_type=None,            #指定填充的类型，支持的有：'solid'等。
                   start_color='ffffffff',    #指定天聪的开始颜色
                   end_color='ff000000'       #指定填充的结束颜色
                   )


#使用：Cell.fill 属性 赋值 openpyxl.styles.PatternFill样式对象 来设置样式填充及渐变色等。
H4=ws['H4']
H4.fill = PatternFill(fill_type='solid',
                    start_color=colors.YELLOW,
                    end_color=colors.YELLOW
                    )



#边框设置 样式对象
#####
#使用：openpyxl.style.Border() 来创建样式对象用于 进行边框设置等

#使用：openpyxl.style.Side(style=None, color=None, border_style=None) 
#来指定用于边框的 border_style样式选项。
#注意:如果您没有指定border_style，其他属性将指定没有效果!
#参数border_style=None: 支持的参数选项有如下：
#{'dashDotDot', 'double', 'mediumDashDot', 'slantDashDot', 'thick', 'dashed', 
#'dashDot', 'mediumDashDotDot', 'mediumDashed', 'medium', 'thin', 'hair', 'dotted'}
help(openpyxl.styles.Border)
help(openpyxl.styles.Side)

border = Border(left=Side(border_style=None, color='FF000000'),        #左边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                right=Side(border_style=None, color='FF000000'),       #右边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                top=Side(border_style=None, color='FF000000'),         #上边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                bottom=Side(border_style=None, color='FF000000'),      #下边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                diagonal=Side(border_style=None, color='FF000000'),    #对角线边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                diagonal_direction=0,                                  #对角线方向
                outline=Side(border_style=None, color='FF000000'),     #外边框线设置，Side类定义 变类型/颜色。'thin'/'thick'
                vertical=Side(border_style=None, color='FF000000'),    #垂直线设置，Side类定义 变类型/颜色。'thin'/'thick'
                horizontal=Side(border_style=None, color='FF000000'),  #水平线设置，Side类定义 变类型/颜色。'thin'/'thick'
                diagonalDown=False,
                start=None,
                end=None
                )

#使用：Cell.border 属性 赋值 openpyxl.styles.Border样式对象 来添加cell单元格/合并的单元的边框属性。
#注意：合并的单元格的行为类似于其他单元格的对象。其值和格式在其左上角的单元格中定义。
ws.merge_cells('A29:F39')                              #合并单元格
left_top_cell=ws['A29']                                #选取合并单元格 的 左上角单元格cell

left_top_cell.value='通过merge_cells()合并的单元格内容'  #设置 合并单元格的值。

thick=Side(border_style='thick',color='ff0000')        #粗线条边框
medium=Side(border_style='medium',color='ff0000')      #中等线条边框
double=Side(border_style='double',color='ff0000')      #双线条边框
thin=Side(border_style='thin',color='ff0000')      #双线条边框

#深入：合并单元格 的 全部边框设置。 官网给出的示例不准确。
#注意：如果要改变所有合并单元的边框，需要对每个单元格都进行边框设置。
left_top_cell.border=Border(left=thick, top=thick, right=thick, bottom=thick)

for row in ws['A29:F39']:
    for cell in row:
        cell.border =  Border(left=medium, top=medium, right=medium, bottom=medium)

ws['A42'].border=Border(left=medium, top=double, right=medium, bottom=double)



#设置对齐 样式对象
#####
#使用：openpyxl.styles.Alignment(horizontal=None, vertical=None, textRotation=0, 
#wrapText=None, shrinkToFit=None, indent=0, relativeIndent=0, justifyLastLine=None, 
#readingOrder=0, text_rotation=None, wrap_text=None, shrink_to_fit=None, mergeCell=None)
#来创建样式对象用于 设置对齐选项 的样式对象
help(openpyxl.styles.Alignment)

alignment = Alignment(horizontal='general',  #水平对齐('centerContinuous', 'general', 'distributed','left', 'fill', 'center', 'justify', 'right')
                      vertical='bottom',     #垂直对齐（'distributed', 'top', 'center', 'justify', 'bottom'）
                      text_rotation=0,       #文字旋转
                      wrap_text=False,       #自动换行
                      shrink_to_fit=False,   #缩小字体填充
                      mergeCell=None,        #合并单元格
                      indent=0               #缩进
                      )

#使用：Cell.alignment 属性 赋值 openpyxl.styles.Alignment样式对象 来设置单元格字体对齐模式。
left_top_cell.alignment=Alignment(horizontal='center', vertical='center')
left_top_cell.font=Font(size=12)                       #设置字体大小


#####
#使用：Cell.number_format 属性来获取/赋值相关的数字格式
number_format='General'


#####
#使用：openpyxl.styles.Protection(locked=True, hidden=False)
#来创建样式对象用于 设置样式的保护选项。(右键、设置单元格样式、保护选项)
protection=Protection(locked=True,
                      hidden=False)



########## #命名样式NameeStyle/样式模板: 
#openpyxl.styles.NamedStyle(), Workbook.add_named_style(), Cell.style;
#注意：
#    命名样式在首次分配给单元时也将自动注册 如ws['A1'].style = NamedStyle(name='namedstyle')
#    注册后，仅使用名称分配样式 如ws['A2'].style=namedstyle

#####
#使用：openpyxl.styles.NamedStyle(self, name='Normal', font=None, **) 类 
#来创建一个命名样式类对象，从而一次将格式应用于许多不同单元格。
help(openpyxl.styles.NamedStyle)

highlight=NamedStyle(name='highlight')      #实例化创建 命名样式
highlight.font=Font(bold=True,size=20)      #设置命名样式的 字体样式
thick=Side(border_style='thick',color=colors.RED)    #设置命名样式的 侧边框样式
highlight.border=Border(left=thick,top=thick,right=thick,bottom=thick)  #设置命名样式的 边框设置

#####
#使用：Workbook.add_named_style(namedStyle)  来将 NamedStyle命名样式 注册到工作簿中。
help(Workbook.add_named_style)
wb.add_named_style(highlight)

#####
#注意：namedStyle命名样式也可以分配给单元格，且首次分配给单元格需要注册，注册后，仅使用名称分配样式即可。
#注意：将命名样式分配给单元后，对该样式的其他更改将 不会影响该单元
ws['A2'].style=highlight
ws['B1'].style='highlight'




#################### 工作表单的页面设置：
#Worksheet.page_setup.orientation / paperSize / fitToHeight / fitToWidth

#Worksheet.print_options.horizontalCentered / verticalCentered
#Worksheet.oddHeader.left.text / size / font / color
#Worksheet.print_title_cols / rows
#Worksheet.print_area

#页面纸张等设置
#####
#使用：#Worksheet.page_setup.orientation / paperSize / fitToHeight / fitToWidth
#等多个属性 来 赋值定义 工作表单的页面设置。
help(ws.page_setup)
dir(ws.page_setup)

ws.page_setup.orientation=ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A4              #设置纸张大小
ws.page_setup.fitToHeight = 0
ws.page_setup.fitToWidth = 1


#编辑打印选项
#####
#使用：Worksheet.print_options.horizontalCentered= True 来设置打印选项 水平居中
#使用：Worksheet.print_options.verticalCentered= True  来设置打印选项 垂直居中
ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = True

#页眉和页脚
####
#使用：Worksheet.odddHeader.left.text/size/font/color 等属性来赋值设置页眉和页脚
#注意：页眉和页脚使用自己的格式语言。
#编写它们时完全支持此操作，但是由于复杂性和嵌套的可能性，在读取它们时仅部分支持。
#左，中/中或右元素支持字体，大小和颜色。
ws.oddHeader.left.text="Page &[Page] of &N"
ws.oddHeader.left.size= 14
ws.oddHeader.left.font="Tahoma,Bold"
ws.oddHeader.left.color="CC3366"

#添加打印标题
####
#使用：Worksheet.print_title_cols/rows 两个属性 来赋值区域 添加打印标题
ws.print_title_cols='A:B'
ws.print_title_rows='1:1'

#设置打印区域
#####
#使用：Worksheet.print_area 属性 赋值添加区域 来设置添加工作表的唯一要打印的部分
ws.print_area='A1:F10'




#################### 工作簿数据存储：Workbook.save()
#保存到文件
#####
#使用：Workbook.save(self, filename) 来将当前工作簿保存到给定的参数filename“文件名”下。
#注意：Workbook.save()操作会覆盖文件，而不会发出警告。
help(Workbook.save)
wb.save(wb_filename)
