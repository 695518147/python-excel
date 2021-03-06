# -*- coding:utf-8 -*-
'''
Author: zhangpeiyu
Date: 2020-08-04 23:56:08
LastEditTime: 2020-08-05 23:57:33
Description: 我不是诗人，所以，只能够把爱你写进程序，当作不可解的密码，作为我一个人知道的秘密。
'''
from openpyxl import Workbook
import os
from openpyxl.styles import Font

# 数据条件格式
from openpyxl.formatting.rule import DataBarRule
rule = DataBarRule(start_type='percentile', start_value=10, end_type='percentile', 
end_value='90',color="FF638EC6", showValue="None", minLength=None, maxLength=None)
rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

'''
根据数据生成excel
'''
def createExcel(header=[], data=[], wb_name = 'wb-name', sheet_name = 'sheet-name'):
    # 创建Excel工作簿
    wb = Workbook()

    for i in wb:
        print(i)
    # 激活当前sheet页
    sheet = wb.active

    # 设置sheet页名字
    sheet.title = sheet_name

    # 添加头部数据
    sheet.append(header)

    # 填充具体的数据
    for item in data:
        sheet.append(item)

    font = Font(name='Calibri',
            size=11,
            color='1874CD',
            bold=True,
            italic=True,
            vertAlign=None,
            underline='none',
            strike=False)
    col = sheet.column_dimensions['A']
    col.font = Font(color="1874CD")
    row = sheet.row_dimensions[1]
    row.font = font
    # 设置行高
    sheet['A1']='行高被设置为 100'
    sheet.row_dimensions[1].height=100
    sheet.column_dimensions['A'].width=100
    

    sheet['A1'].font = font

    print(sheet.print_area)
    
    sheet.conditional_formatting.add('A1:F40', rule)

    filePath = os.path.join(os.path.dirname(__file__), wb_name + ".xlsx")
    wb.save(filePath)
    print(u'最大行数：', sheet.max_row)
    print(u'最大列数：', sheet.max_column)
    print(u'生成的Excel路径是：', filePath)
